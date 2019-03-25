#! "C:\Program Files (x86)\Python37-32\python.exe"

import openpyxl

from openpyxl import load_workbook

file = "MOCK_DATA.xlsx"

#load the work book
wb_obj = load_workbook(filename = file)

wsheet = wb_obj['MOCK_DATA']

#store data in a dictionary for convenience
dataDict = {}
value = []

for key, *values in wsheet.iter_rows(min_row=2):
    dataDict[key.value] = [v.value for v in values]

# print ("Contents of your spreadsheet :\n",dataDict)

#prompt user for input(ID)
userInput = input("Please enter an ID of a person to find a his/her details: ")

#print details of a person associated with above ID

if (int(userInput)) in dataDict:

    print("Details requested for the ID:",int(userInput))

    print(dataDict.get(int(userInput)))
else:
    print("The ID",userInput,"is not found. Please enter a valid id")
